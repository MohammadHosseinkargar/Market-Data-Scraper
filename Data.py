import requests
import pandas as pd
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
from io import StringIO
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# دریافت ورودی کاربر برای انتخاب روز
user_input = input("Enter 'T' for today's data or 'Y' for yesterday's data: ").strip().upper()

# تنظیم تاریخ بر اساس ورودی کاربر
if user_input == 'T':
    date = datetime.now()
elif user_input == 'Y':
    date = datetime.now() - timedelta(1)
else:
    raise ValueError("Invalid input! Please enter 'T' for today's data or 'Y' for yesterday's data.")

date_str = date.strftime('%Y-%m-%d')

# لینک با تاریخ انتخاب شده
url = f"https://hupx.hu/en/market-data/id/market-data?date={date_str}"

# درخواست به صفحه وب
response = requests.get(url)
response.raise_for_status()  # بررسی موفقیت درخواست

# پارس کردن محتوای HTML
soup = BeautifulSoup(response.content, 'html.parser')

# پیدا کردن جداول مرتبط با تب‌ها
hours_tab = soup.find('div', {'id': 'hours'})
quarters_tab = soup.find('div', {'id': 'quarters'})

if hours_tab is None or quarters_tab is None:
    raise ValueError("حداقل یکی از تب‌های مورد نیاز پیدا نشد.")

# استخراج جداول از تب‌ها
hours_table = hours_tab.find('table')
quarters_table = quarters_tab.find('table')

if hours_table is None or quarters_table is None:
    raise ValueError("حداقل یکی از جداول مورد نیاز پیدا نشد.")

# استخراج جداول به دیتافریم‌های pandas
df_hours = pd.read_html(StringIO(str(hours_table)))[0]
df_quartally = pd.read_html(StringIO(str(quarters_table)))[0]

# تنظیم فرمت نام فایل‌ها با تاریخ و ساعت
current_time = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
hours_file = f'market_data_hours_{current_time}.xlsx'
quarters_file = f'market_data_quarters_{current_time}.xlsx'

# تابعی برای تنظیم عرض ستون‌ها
def adjust_column_width(writer, sheet_name, dataframe):
    worksheet = writer.sheets[sheet_name]
    for col_idx, col in enumerate(dataframe.columns, 1):
        max_length = max(dataframe[col].astype(str).map(len).max(), len(col)) + 2
        worksheet.column_dimensions[get_column_letter(col_idx)].width = max_length

# تابعی برای تنظیم وسط‌چین بودن مقادیر در ستون‌ها و ردیف‌ها
def set_cell_alignment(worksheet):
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

# ذخیره کردن داده‌های hours به فایل اکسل و تنظیم عرض ستون‌ها
with pd.ExcelWriter(hours_file, engine='openpyxl') as writer:
    df_hours.to_excel(writer, sheet_name='hours', index=False)
    adjust_column_width(writer, 'hours', df_hours)
    worksheet = writer.sheets['hours']
    set_cell_alignment(worksheet)

# ذخیره کردن داده‌های quarters به فایل اکسل و تنظیم عرض ستون‌ها
with pd.ExcelWriter(quarters_file, engine='openpyxl') as writer:
    df_quartally.to_excel(writer, sheet_name='quartally', index=False)
    adjust_column_width(writer, 'quartally', df_quartally)
    worksheet = writer.sheets['quartally']
    set_cell_alignment(worksheet)

print(f"Hours data has been successfully saved to {hours_file}")
print(f"Quarterly data has been successfully saved to {quarters_file}")
